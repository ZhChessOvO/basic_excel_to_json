import tkinter
from tkinter import *
from tkinter import messagebox
import openpyxl
import json


def get_file(name):
    try:
        name = name + '.xlsx'
        file = openpyxl.load_workbook(name)
        sheet_name = file.sheetnames[0]

        sheet = file[sheet_name]

        # print(sheet.max_row, sheet.max_column)
        # print(sheet['A32'].value)

        col_num = sheet.max_column  # 一共有几行
        # print(sheet.cell(row=1,column=1))

        content = {}
        for col in range(1, col_num + 1):
            line = 1
            col_name = sheet.cell(row=line, column=col).value
            content[col_name] = []
            line = 2
            while True:
                curr_data = sheet.cell(row=line, column=col).value
                if curr_data is not None:
                    line += 1
                    content[col_name].append(curr_data)
                else:
                    break

        content_text = json.dumps(content, indent=4, ensure_ascii=False)
        json_text.insert(INSERT, content_text)

        file.close()

    except IOError:
        messagebox.showinfo('错误提示', name + '不存在，请检查是否在同目录下')
    return


def load_file(name):  # 导出
    try:
        file_name = name + '.json'
        file = open(file_name, 'w')
        file.write(json_text.get(1.0, tkinter.END))
        file.close()
        messagebox.showinfo('保存成功！', file_name+'已成功保存在同目录下')

    except IOError:
        messagebox.showinfo('错误提示', '出现未知错误，请检查文件名是否有非法字符或手动复制文本框内内容！')
    return


root = Tk()
root.title('单层json生成器 ver1.0 by ChessZH')

Label(root, text='').grid(row=0)  # 控制格式的空行
Label(root, text='excel文件名:').grid(row=1, column=1)
file_name = StringVar()
out_name = StringVar()
Entry(root, textvariable=file_name, width=20).grid(row=1, column=2)
Label(root, text='.xlsx     ').grid(row=1, column=3)
Button(text="导入文件", command=lambda: get_file(file_name.get())).grid(row=1, column=4)
Label(root, text=' ').grid(row=1, column=5)
Label(root, text='').grid(row=2)  # 控制格式的空行

# 文本框主体
json_text = Text(root, width=50, height=20)
json_text.grid(row=3, column=1, columnspan=4)

# Button(text='GO', command=lambda: clickButton(file_name.get(), grade0.get())).grid(row=3, column=2, sticky=E)
Label(root, text='').grid(row=4)  # 控制格式的空行

Label(root, text='导出为').grid(row=5, column=1)
Entry(root, textvariable=out_name, width=20).grid(row=5, column=2)
Label(root, text='.json     ').grid(row=5, column=3)
Button(text="生成！", command=lambda: load_file(out_name.get())).grid(row=5, column=4)

Label(root, text='').grid(row=6)  # 控制格式的空行

root.mainloop()
